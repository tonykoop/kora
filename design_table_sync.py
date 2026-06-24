#!/usr/bin/env python3
"""
CSV <-> XLSX sync for kora-design-table.xlsx.

The XLSX is the editable / SolidWorks-facing workbook; the CSV is the
diffable, corruption-proof mirror that lives in git history. Keep them in sync:

    python design_table_sync.py pull     # XLSX -> CSV  (after editing in Excel)
    python design_table_sync.py push     # CSV  -> XLSX (rebuild workbook)
    python design_table_sync.py check    # report whether they match

Commit the CSV. The XLSX is handled by Git LFS (see .gitattributes) so binary
churn never corrupts the index. If the XLSX ever goes bad again, run `push`
to regenerate it from the CSV, or `git checkout HEAD -- kora-design-table.xlsx`.

Round-trip preserves cell VALUES of the 'Kora' sheet, not Excel formatting.
"""
import sys, csv, os
import openpyxl

XLSX  = "kora-design-table.xlsx"
CSV   = "kora-design-table.csv"
SHEET = "Kora"


def _num(v):
    if v is None:
        return ""
    return v


def pull():
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    with open(CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            w.writerow(["" if c is None else c for c in row])
    print(f"pull: {XLSX}[{ws.title}] -> {CSV}")


def _coerce(s):
    if s == "":
        return None
    try:
        i = int(s)
        return i
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        return s


def push():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    with open(CSV, newline="", encoding="utf-8") as f:
        for row in csv.reader(f):
            ws.append([_coerce(c) for c in row])
    wb.save(XLSX)
    print(f"push: {CSV} -> {XLSX}[{SHEET}]")


def check():
    if not (os.path.exists(XLSX) and os.path.exists(CSV)):
        print("check: missing file"); return
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    x = [["" if c is None else str(c) for c in r] for r in ws.iter_rows(values_only=True)]
    with open(CSV, newline="", encoding="utf-8") as f:
        c = [[str(v) for v in row] for row in csv.reader(f)]
    print("check: MATCH" if x == c else "check: DIFFER (run pull or push)")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "check"
    {"pull": pull, "push": push, "check": check}.get(cmd, check)()
